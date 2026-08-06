import sys
import os
# Tambahkan path root agar bisa import utils jika folder terpisah
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import io
import base64

import tempfile
import datetime

import requests
import json
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS

# Media Processing
import threading
import subprocess
import tempfile

# Excel
from openpyxl import Workbook

# PDF & Images
import fitz  # PyMuPDF
from pypdf import PdfReader, PdfWriter
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import black, white, gray, Color
from reportlab.pdfbase.pdfmetrics import stringWidth

# Utils (Asumsi Anda punya file utils/db.py dan utils/storage.py)
# Jika tidak, Anda bisa menggabungkan kode koneksi DB/R2 di sini langsung.
from utils.db import get_members_collection
from utils.storage import upload_bytes_to_r2, get_s3_client
from utils.helpers import flatten_object, month_to_roman, draw_wrapped_text, format_date_indo

app = Flask(__name__)
CORS(app)

import jwt

JWT_SECRET = os.environ.get("JWT_SECRET")

@app.before_request
def verify_jwt():
    if request.path == "/" and request.method == "GET":
        return None
    
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return jsonify({"error": "Missing or invalid token"}), 401
    
    token = auth_header.split(" ")[1]
    if not JWT_SECRET:
        print("WARNING: JWT_SECRET is not set. Requests are not authenticated securely.")
        return None
        
    try:
        jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401

# --- HELPER FUNCTIONS ---


def draw_signature_box(c, x, y_top, w, h, p_h,
                        sig_name='', sig_as='', sig_img=None, overlap=False):
    """
    Layout kotak tanda tangan:
        ┌─────────────────┐
        │  Jabatan (atas) │  ← 15% h, font min 12pt
        ├─────────────────┤
        │  [TTD / QR]     │  ← gambar besar, ujung bawahnya
        │    ↕ overlap    │    menumpangi area nama (~40%)
        ├─────────────────┤  ← garis tipis
        │  Nama (bawah)   │  ← 20% h, font min 12pt
        └─────────────────┘
    Gambar digambar TERAKHIR sehingga secara visual
    "tinta tanda tangan" berada di atas teks nama.
    sig_img : ImageReader-compatible, atau None
    """
    role_h   = h * 0.15   # jabatan di atas
    name_h   = h * 0.20   # nama di bawah
    img_h    = h - role_h - name_h   # sisa → gambar

    # Koordinat ReportLab (bottom-left origin)
    y_bot    = p_h - (y_top + h)
    name_y0  = y_bot               # bawah area nama
    name_y1  = y_bot + name_h      # batas atas nama / batas bawah garis
    img_y0   = name_y1             # bawah area gambar (tanpa overlap)
    img_y1   = img_y0 + img_h      # atas area gambar / batas bawah jabatan
    role_y0  = img_y1

    c.saveState()

    # ── 1. Jabatan teks (atas) ─────────────────────────────────
    if sig_as:
        role_fs = max(12, min(14, role_h * 0.60))
        c.setFont('Helvetica', role_fs)
        c.setFillColor(Color(0.2, 0.2, 0.2))
        role_text_y = role_y0 + (role_h - role_fs) / 2
        c.drawCentredString(x + w / 2, role_text_y, sig_as[:50])

    # ── 2. Garis atas (antara jabatan & gambar) ────────────────
    c.setStrokeColor(Color(0.5, 0.5, 0.5))
    c.setLineWidth(0.3)
    c.line(x, img_y1, x + w, img_y1)

    # ── 3. Garis bawah (antara gambar & nama) ─────────────────
    c.setStrokeColor(Color(0.5, 0.5, 0.5))
    c.setLineWidth(0.5)
    c.line(x, name_y1, x + w, name_y1)

    # ── 4. Nama teks (bawah, di bawah garis) ──────────────────
    if sig_name:
        name_fs = max(12, min(14, name_h * 0.60))
        c.setFont('Helvetica-Bold', name_fs)
        c.setFillColor(Color(0.1, 0.1, 0.1))
        name_text_y = name_y0 + (name_h - name_fs) / 2
        c.drawCentredString(x + w / 2, name_text_y, sig_name[:40])

    # ── 5. Gambar TTD / QR (digambar TERAKHIR → di atas nama) ──
    if sig_img is not None:
        if overlap:
            # Wet signature: gambar tumpang-tindih ke area nama (50% name_h)
            overlap_px  = name_h * 0.50
            draw_y      = img_y0 - overlap_px
            draw_h      = img_h + overlap_px
        else:
            # QR: tetap di dalam area gambar, tidak tumpang tindih
            draw_y  = img_y0
            draw_h  = img_h
        side     = min(w * 0.92, draw_h)
        draw_x   = x + (w - side) / 2
        center_y = draw_y + (draw_h - side) / 2
        c.drawImage(sig_img, draw_x, center_y,
                    width=side, height=side,
                    mask='auto', preserveAspectRatio=False)

    c.restoreState()

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "service": "Himatika Python Worker",
        "status": "Online",
        "timestamp": datetime.datetime.now().isoformat(),
        "endpoints": {
            "excel_export": "/api/sheet/export",
            "excel_import": "/api/sheet/import",
            "pdf_signature": "/api/sign/process",
            "pdf_search_text": "/api/pdf/search-text",
            "activiness_letter": "/api/pdf/activiness-letter",
            "ticket_generator": "/api/pdf/ticket",
            "video_compress": "/api/media/compress-video"
        }
    })
# --- ENDPOINTS ---

# 0. EXCEL IMPORT
# 0. EXCEL IMPORT
from openpyxl import load_workbook
import zxingcpp
from PIL import Image

@app.route('/api/pdf/scan-qr', methods=['POST'])
def scan_qr():
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Save to temp
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            file.save(temp_pdf.name)
            temp_pdf_path = temp_pdf.name

        doc = fitz.open(temp_pdf_path)
        found_data = None
        
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            
            # Improve resolution (3x zoom)
            zoom = 3.0
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert PyMuPDF Pixmap to PIL Image
            if pix.n >= 4:
                mode = "RGBA"
            elif pix.n == 3:
                mode = "RGB"
            else:
                mode = "L"
            
            img_pil = Image.frombytes(mode, [pix.width, pix.height], pix.samples)

            # Detect attempts with zxing-cpp using Pillow Images
            attempts = [img_pil]
            
            # 2. Grayscale
            if mode != 'L':
                img_gray = img_pil.convert('L')
                attempts.append(img_gray)
            else:
                img_gray = img_pil
            
            # 3. Binary Threshold (Simulation of cv2.threshold)
            # Threshold = 128
            img_binary = img_gray.point(lambda p: 255 if p > 128 else 0, mode='1')
            attempts.append(img_binary)

            # 4. Darker Threshold
            img_binary_dark = img_gray.point(lambda p: 255 if p > 100 else 0, mode='1')
            attempts.append(img_binary_dark)
            
            for img_check in attempts:
                try:
                    results = zxingcpp.read_barcodes(img_check)
                    if results:
                        for result in results:
                            if result.format == zxingcpp.BarcodeFormat.QRCode:
                                found_data = result.text
                                break
                    if found_data:
                        break
                except Exception as e:
                    # Continue if decoding fails for one variant
                    continue
            
            if found_data:
                break
                
        doc.close()
        os.remove(temp_pdf_path)

        if found_data:
             return jsonify({'status': 'found', 'data': found_data})
        else:
             return jsonify({'status': 'not_found'})

    except Exception as e:
        print(f"Error scanning QR: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet/import', methods=['POST'])
def import_generic_sheet():
    try:
        if 'file' not in request.files:
             return jsonify({"error": "No file part"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active # Default to first sheet or active one
        
        data = []
        headers = []
        
        # Read headers (Row 1)
        header_row_gen = ws.iter_rows(min_row=1, max_row=1)
        try:
            header_row = next(header_row_gen)
            for cell in header_row:
                headers.append(cell.value)
        except StopIteration:
            pass # Empty sheet
            
        # Read data (Row 2 onwards)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
                
            row_data = {}
            for i, cell_value in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    if header is None: continue

                    # Handle nested keys (e.g. "user.name")
                    keys = str(header).split('.')
                    current_ref = row_data
                    
                    for k_idx, key in enumerate(keys):
                        if k_idx == len(keys) - 1:
                            # Last key, assign value
                            current_ref[key] = cell_value
                        else:
                            # Create nested dict if not exists
                            if key not in current_ref:
                                current_ref[key] = {}
                            current_ref = current_ref[key]
                            
            data.append(row_data)
            
        return jsonify({
            "success": True,
            "data": data
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# 0.5 PDF TEXT SEARCH (For Auto Signature)
@app.route('/api/pdf/search-text', methods=['POST'])
def search_text_pdf():
    try:
        body = request.json
        pdf_url = body.get('pdf')
        search_text = body.get('text')
        
        if not pdf_url or not search_text:
             return jsonify({"error": "Missing pdf url or search text"}), 400
             
        # Download PDF
        response = requests.get(pdf_url)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch PDF"}), 400
            
        pdf_bytes = response.content
        
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        matches = []
        
        for page_num, page in enumerate(doc):
            # search for text
            text_instances = page.search_for(search_text)
            
            for inst in text_instances:
                # inst is Rect(x0, y0, x1, y1)
                # Frontend expects: { page, x, y, width, height }
                # NOTE: PyMuPDF coordinates are Top-Left (0,0). 
                # PDF.js also uses Top-Left usually, but sometimes Bottom-Left depending on view.
                # app/pages/signatures/[id].vue logic:
                # const y = viewport.height - transform[5]; // Konversi Y dari bawah ke atas if using PDF.js raw
                # But here we return coordinates. Let's return standard Top-Left X, Y.
                
                matches.append({
                    "page": page_num + 1, # 1-based index for frontend consistency
                    "x": inst.x0,
                    "y": inst.y0, # Top
                    "width": inst.width,
                    "height": inst.height,
                    "text": search_text
                })
                
        doc.close()
        
        return jsonify(matches)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# 1. GENERIC EXCEL EXPORT
@app.route('/api/sheet/export', methods=['POST'])
def export_generic_sheet():
    try:
        body = request.json
        complex_data = body.get('data', [])
        title = body.get('title', 'Export')
        headers_input = body.get('headers', None)

        wb = Workbook()
        ws = wb.active
        ws.title = title[:30] # Excel sheet limit

        if not complex_data:
            return jsonify({"message": "No data"}), 400

        if headers_input:
            headers = headers_input
        else:
            all_headers = set()
            for item in complex_data:
                flat = flatten_object(item)
                all_headers.update(flat.keys())
            headers = list(all_headers)
            
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for item in complex_data:
            flat = flatten_object(item)
            row = [str(flat.get(h, "")) if isinstance(flat.get(h, ""), (list, dict)) else flat.get(h, "") for h in headers]
            ws.append(row)

        # Style data cells
        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_alignment

        # Auto adjust column width
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 60) # Max width to avoid overly wide columns

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{title}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 2. PDF SIGNATURE PROCESSOR (Auto-Detect / Manual Location)
@app.route('/api/sign/process', methods=['POST'])
def process_sign_overlay():
    try:
        body = request.json
        pdf_url = body.get('pdf')
        output_path = body.get('outputBlobPath')
        qr_value = body.get('qrValue')
        signer_name = body.get('signerName', '')  # nama member penandatangan
        signer_as   = body.get('signerAs', '')    # jabatan penandatangan
        
        # Opsi: Manual locations ATAU Search Text
        manual_locations = body.get('locations', [])

        if not all([pdf_url, output_path, qr_value]):
            return jsonify({"error": "Missing parameters"}), 400

        # Download PDF
        response = requests.get(pdf_url)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch PDF"}), 400
        
        pdf_bytes = response.content
        input_pdf_stream = io.BytesIO(pdf_bytes)

        target_locations = manual_locations
        if not target_locations:
            return jsonify({"error": "No signature location found."}), 400

        # Generate QR Code
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, border=1, box_size=10)
        qr.add_data(qr_value)
        qr.make(fit=True)
        qr_pil = qr.make_image(fill_color="black", back_color="white")
        
        qr_bytes = io.BytesIO()
        qr_pil.save(qr_bytes, format='PNG')
        qr_bytes.seek(0)
        qr_img = ImageReader(qr_bytes)

        reader = PdfReader(input_pdf_stream)
        writer = PdfWriter()

        # Group Locations
        locs_by_page = {}
        for loc in target_locations:
            p = int(loc.get('page', 1)) - 1
            if p not in locs_by_page: locs_by_page[p] = []
            locs_by_page[p].append(loc)

        for i, page in enumerate(reader.pages):
            if i in locs_by_page:
                p_w = float(page.mediabox.width)
                p_h = float(page.mediabox.height)
                
                packet = io.BytesIO()
                can = canvas.Canvas(packet, pagesize=(p_w, p_h))

                for loc in locs_by_page[i]:
                    x = float(loc.get('x', 0))
                    y_top = float(loc.get('y', 0))
                    w = float(loc.get('width', 100))
                    h = float(loc.get('height', 100))

                    # Hanya gambar QR, abaikan teks dan garis
                    y_bot_qr = p_h - (y_top + h)
                    # Use a centered square for the QR code to ensure it's not distorted and centered
                    side = min(w, h)
                    draw_x = x + (w - side) / 2
                    draw_y = y_bot_qr + (h - side) / 2
                    can.drawImage(qr_img, draw_x, draw_y, width=side, height=side, mask='auto', preserveAspectRatio=True)

                can.save()
                packet.seek(0)
                page.merge_page(PdfReader(packet).pages[0])
                packet.close()

            writer.add_page(page)

        output_buffer = io.BytesIO()
        writer.write(output_buffer)
        output_buffer.seek(0)
        
        public_url = upload_bytes_to_r2(output_buffer.getvalue(), "application/pdf", output_path)

        return jsonify({
            "success": True, 
            "data": public_url
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        if 'input_pdf_stream' in locals(): input_pdf_stream.close()
        if 'qr_bytes' in locals(): qr_bytes.close()
        if 'output_buffer' in locals(): output_buffer.close()


# 3. GENERATE ACTIVINESS LETTER (Dengan Return Lokasi Tanda Tangan)
@app.route('/api/pdf/activiness-letter', methods=['POST'])
def generate_activiness_letter():
    try:
        body = request.json
        member_data = body.get('member')
        point_data = body.get('point')
        chairman_data = body.get('chairman')
        secretary_data = body.get('secretary')
        doc_number = body.get('docNumber')
        period = body.get('period')
        config_data = body.get('config', {})
        
        if not all([member_data, point_data, chairman_data, secretary_data, doc_number, period]):
            return jsonify({"error": "Incomplete data"}), 400

        # Strict validation for NIM
        if not member_data.get('NIM') or str(member_data.get('NIM')).strip() == '':
            return jsonify({"error": "NIM tidak valid"}), 400

        # Setup PDF
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        margin = 40
        def draw_header(c, y_pos):
            # Base Dir untuk Assets
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            
            def draw_logo_boxed(path, x, y, max_w, max_h, align='left'):
                try:
                    if not os.path.exists(path):
                        return
                    img = ImageReader(path)
                    iw, ih = img.getSize()
                    aspect = iw / float(ih)
                    
                    # Calculate new dimensions fitting in box
                    if aspect > 1: # Wide
                        w = min(iw, max_w)
                        h = w / aspect
                        if h > max_h:
                            h = max_h
                            w = h * aspect
                    else: # Tall or Square
                        h = min(ih, max_h)
                        w = h * aspect
                        if w > max_w:
                            w = max_w
                            h = w / aspect
                    
                    # Draw
                    draw_x = x
                    if align == 'right':
                        draw_x = x + (max_w - w) # Align right within the box (or just x if x is right edge... logic below)
                    elif align == 'center':
                        draw_x = x + (max_w - w)/2
                    
                    # Note: x passed to this function for 'right' logo is usually the left edge of the right area
                    # But in previous code: width - margin - 60. So that is the LEFT edge of the 60px box.
                    
                    c.drawImage(img, draw_x, y + (max_h - h)/2, width=w, height=h, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo {path}: {e}")

            # Logo Left (ITSNU)
            logo_itsnu_path = os.path.join(base_dir, 'assets', 'img', 'itsnu-logo.png')
            draw_logo_boxed(logo_itsnu_path, margin, y_pos - 15, 80, 80, align='left')

            # Logo Right (HIMATIKA)
            logo_hima_path = os.path.join(base_dir, 'assets', 'img', 'logo.png')
            # Right logo position: width - margin - 80 is the LEFT x of the right logo box
            draw_logo_boxed(logo_hima_path, width - margin - 80, y_pos - 15, 80, 80, align='right')

            # Center Text
            text_y = y_pos + 50
            c.setFont("Times-Bold", 12)
            c.drawCentredString(width/2, text_y, config_data.get('name', 'Himatika'))
            c.setFont("Times-Bold", 12)
            c.drawCentredString(width/2, text_y - 12, "FAKULTAS SAINS DAN TEKNOLOGI")
            c.drawCentredString(width/2, text_y - 24, "INSTITUT TEKNOLOGI DAN SAINS NU PEKALONGAN")
            c.setFont("Times-Bold", 12)
            c.drawCentredString(width/2, text_y - 36, period)
            
            # Address & Contact
            c.setFont("Times-Italic", 9)
            address_str = config_data.get('address', "Jl. Kampus No. 1, Gedung PKM Lt. 2")
            phone_str = config_data.get('phone', '+62 812-3456-7890')
            email_str = config_data.get('email', 'himatika@unvc.ac.id')
            
            contact_line1 = f"Sekretariat: {address_str}"
            contact_line2 = f"Narahubung: {phone_str} | Surel: {email_str}"

            c.drawCentredString(width/2, text_y - 45, contact_line1)
            c.drawCentredString(width/2, text_y - 56, contact_line2)

            # Double Line
            line_y = y_pos - 15
            c.setLineWidth(2)
            c.line(margin, line_y, width - margin, line_y)
            c.setLineWidth(0.5)
            c.line(margin, line_y - 2, width - margin, line_y - 2)
            
            return line_y - 20

        # --- PAGE 1: SURAT ---
        header_bottom = draw_header(c, height - 100)

        # Judul
        title_y = header_bottom
        c.setFont("Times-Bold", 12)
        c.drawCentredString(width/2, title_y, "Surat Keterangan Aktif")
        
        c.setFont("Times-Bold", 12) 
        # Manual underline
        org_text = "Himpunan Mahasiswa Informatika"
        text_w = stringWidth(org_text, "Times-Bold", 12)
        c.drawCentredString(width/2, title_y - 18, org_text)
        c.setLineWidth(1)
        c.line((width/2) - (text_w/2), title_y - 20, (width/2) + (text_w/2), title_y - 20)
        
        c.setFont("Times-Roman", 12)
        c.drawCentredString(width/2, title_y - 35, doc_number)

        # Body
        body_y = title_y - 70
        c.setFont("Times-Roman", 12)
        c.drawString(margin, body_y, "Yang bertanda tangan di bawah ini :")
        
        def draw_kv(lbl, val, y):
            c.drawString(margin, y, lbl)
            c.drawString(margin + 100, y, ": " + str(val))
            return y - 18

        body_y -= 25
        body_y = draw_kv("Nama", chairman_data.get('fullName', '-'), body_y)
        body_y = draw_kv("NIM", chairman_data.get('NIM', '-'), body_y)
        body_y = draw_kv("Jabatan", "Ketua Umum", body_y)

        body_y -= 15
        c.drawString(margin, body_y, "Menyatakan dengan sesungguhnya bahwa :")
        body_y -= 25
        
        body_y = draw_kv("Nama", member_data.get('fullName', '-'), body_y)
        body_y = draw_kv("NIM", member_data.get('NIM', '-'), body_y)
        member_class = member_data.get('class', '-')
        body_y = draw_kv("Kelas", member_class, body_y)
        body_y = draw_kv("Semester", point_data.get('semester', '-'), body_y)

        body_y -= 20
        text = f"Adalah mahasiswa yang benar - benar aktif dalam Himpunan Mahasiswa Informatika (HIMATIKA) ITSNU Pekalongan periode {period}."
        body_y = draw_wrapped_text(c, text, margin, body_y, width - 2*margin, "Times-Roman", 12)

        body_y -= 15
        text2 = "Demikian surat keterangan keaktifan mahasiswa ini dibuat sebagaimana mestinya."
        body_y = draw_wrapped_text(c, text2, margin, body_y, width - 2*margin, "Times-Roman", 12)

        if isinstance(period, str) and len(period) == 4 and period.isdigit():
             # If period is just year, try to parse it or just use it? 
             # Actually `period` is string like "2023/2024". Just verify context.
             pass

        footer_y = body_y - 30
        date_obj = datetime.datetime.now()
        date_str = format_date_indo(date_obj)
        c.drawRightString(width - margin, footer_y, f"Pekalongan, {date_str}")
        
        footer_y -= 20
        c.setFont("Times-Bold", 12)
        c.drawCentredString(width/2, footer_y, "HIMPUNAN MAHASISWA INFORMATIKA")
        c.drawCentredString(width/2, footer_y - 12, "INSTITUT TEKNOLOGI DAN SAINS NAHDLATUL ULAMA")
        c.drawCentredString(width/2, footer_y - 24, "PEKALONGAN")

        footer_y -= 50
        c.setFont("Times-Bold", 12)
        c.drawCentredString(width/2, footer_y, "Mengetahui")
        
        sig_y_start = footer_y - 30
        left_x = width / 4
        right_x = (width * 3) / 4
        
        # Jabatan
        c.setFont("Times-Bold", 12)
        c.drawCentredString(left_x, sig_y_start, "Ketua Umum")
        c.drawCentredString(right_x, sig_y_start, "Sekretaris Umum")
        
        # Signature Placeholders (Grey)
        sig_space_y = sig_y_start - 40
        c.setFillColor(Color(0.7, 0.7, 0.7))
        c.setFont("Courier", 9)
        c.setFillColor(black)

        # Nama
        name_y = sig_y_start - 80
        c.setFont("Times-Bold", 12)
        c.drawCentredString(left_x, name_y, chairman_data.get('fullName', ''))
        c.drawCentredString(right_x, name_y, secretary_data.get('fullName', ''))
        
        # Underline Nama
        c.setLineWidth(0.5)
        # Left Name Underline
        nw_l = stringWidth(chairman_data.get('fullName', ''), "Times-Bold", 12)
        c.line(left_x - (nw_l/2), name_y - 2, left_x + (nw_l/2), name_y - 2)
        # Right Name Underline
        nw_r = stringWidth(secretary_data.get('fullName', ''), "Times-Bold", 12)
        c.line(right_x - (nw_r/2), name_y - 2, right_x + (nw_r/2), name_y - 2)

        c.setFont("Times-Roman", 12)
        c.drawCentredString(left_x, name_y - 14, str(chairman_data.get('NIM', '')))
        c.drawCentredString(right_x, name_y - 14, str(secretary_data.get('NIM', '')))

        # Footer Note
        note_y = 40
        c.setFont("Times-Italic", 7)
        c.drawString(margin, note_y, f"*Surat ini dibuat dengan menggunakan sistem Informasi Himpunan Mahasiswa Informatika (HIMATIKA) ITSNU Pekalongan")
        c.drawString(margin, note_y - 8, f"*Untuk verifikasi keaslian surat ini, silakan kunjungi: {os.getenv('PUBLIC_URL')}/verify/scan")
        c.drawString(margin, note_y - 16, f"dan ditandatangani secara elektronik. Surat ini sah dan berlaku sebagai bukti keaktifan mahasiswa dalam organisasi.")

        # --- SIGNATURE CALCULATIONS ---
        qr_size = 60
        # Signature is roughly centered on sig_space_y
        qr_bottom_y = sig_space_y - (qr_size / 2)
        fe_y = height - (qr_bottom_y + qr_size) # Top-Left System

        sig_locations = [
            {
                "page": 1,
                "role": "Chairman",
                "nim": str(chairman_data.get('NIM')),
                "x": left_x - (qr_size/2),
                "y": fe_y,
                "width": qr_size,
                "height": qr_size
            },
            {
                "page": 1,
                "role": "Secretary",
                "nim": str(secretary_data.get('NIM')),
                "x": right_x - (qr_size/2),
                "y": fe_y,
                "width": qr_size,
                "height": qr_size
            }
        ]
        
        c.showPage()

        # --- PAGE 2+: LAMPIRAN (Summary Cards + Detail Tables) ---
        activities_details = body.get('activitiesDetails', {})
        activities = point_data.get('activities', {})
        agendas_count = activities.get('agendas', {})

        # Hitung total poin dari data detail yang sebenarnya
        def sum_points(items):
            return sum(item.get('point', 0) for item in items) if items else 0

        computed_total = (
            sum_points(activities_details.get('committees', []))
            + sum_points(activities_details.get('participants', []))
            + sum_points(activities_details.get('projects', []))
            + sum_points(activities_details.get('aspirations', []))
        )
        # Gunakan total dari data detail jika ada, fallback ke point_data
        total_point = computed_total if computed_total > 0 else point_data.get('point', 0)

        # Helper: format date from ISO string
        def fmt_date(d):
            if not d:
                return "-"
            try:
                if isinstance(d, str):
                    dt = datetime.datetime.fromisoformat(d.replace('Z', '+00:00'))
                else:
                    dt = d
                return dt.strftime("%d %b %Y")
            except Exception:
                return str(d)[:10] if d else "-"

        # Batas bawah halaman (untuk page break otomatis)
        page_bottom = 60

        def start_new_page():
            """Buat halaman baru dengan header."""
            c.showPage()
            return draw_header(c, height - 100)

        header_bottom = draw_header(c, height - 100)
        curr_y = header_bottom - 10

        # ── Judul Lampiran ──
        c.setFont("Times-Bold", 14)
        c.drawCentredString(width / 2, curr_y, "LAMPIRAN")
        curr_y -= 16
        c.setFont("Times-Roman", 11)
        subtitle = f"Rincian Keaktifan Mahasiswa Semester {point_data.get('semester', '-')}"
        c.drawCentredString(width / 2, curr_y, subtitle)
        curr_y -= 10
        c.setLineWidth(0.5)
        c.line(margin, curr_y, width - margin, curr_y)
        curr_y -= 20

        # ── SUMMARY CARDS ──
        card_data = [
            ("Kepanitiaan", len(activities_details.get('committees', []))),
            ("Kepesertaan", len(activities_details.get('participants', []))),
            ("Proyek", len(activities_details.get('projects', []))),
            ("Aspirasi", len(activities_details.get('aspirations', []))),
        ]
        num_cards = len(card_data)
        usable_w = width - 2 * margin
        card_gap = 8
        card_w = (usable_w - (num_cards - 1) * card_gap) / num_cards
        card_h = 50
        card_y_top = curr_y  # top of the cards (ReportLab bottom-left)

        for idx, (label, count) in enumerate(card_data):
            cx = margin + idx * (card_w + card_gap)
            cy = card_y_top - card_h  # bottom of card

            # Card background
            c.saveState()
            c.setFillColor(Color(0.95, 0.95, 0.97))
            c.setStrokeColor(Color(0.80, 0.80, 0.85))
            c.setLineWidth(0.5)
            c.roundRect(cx, cy, card_w, card_h, 4, fill=1, stroke=1)

            # Count number (big, centered)
            c.setFillColor(Color(0.15, 0.15, 0.15))
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(cx + card_w / 2, cy + card_h - 22, str(count))

            # Label (small, centered, below number)
            c.setFillColor(Color(0.4, 0.4, 0.4))
            c.setFont("Helvetica", 8)
            c.drawCentredString(cx + card_w / 2, cy + 8, label)
            c.restoreState()

        curr_y = card_y_top - card_h - 12

        # Total Point Badge (centered below cards)
        tp_text = f"Total Poin: {total_point}"
        tp_w = stringWidth(tp_text, "Helvetica-Bold", 11) + 24
        tp_h = 22
        tp_x = (width - tp_w) / 2
        tp_y = curr_y - tp_h

        c.saveState()
        c.setFillColor(Color(0.20, 0.30, 0.55))
        c.roundRect(tp_x, tp_y, tp_w, tp_h, 4, fill=1, stroke=0)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width / 2, tp_y + 6, tp_text)
        c.restoreState()

        curr_y = tp_y - 25

        # ── DETAIL TABLES ──
        # Column layout: No | Nama Kegiatan | Tanggal | Poin
        col_no_x = margin
        col_name_x = margin + 30
        col_date_x = width - margin - 130
        col_point_x = width - margin - 40
        table_right = width - margin
        row_h = 16

        def draw_table_header(y):
            """Draw the table column headers and top/bottom lines."""
            c.setFont("Times-Bold", 10)
            c.setFillColor(Color(0.15, 0.15, 0.15))
            # Top line
            c.setLineWidth(0.8)
            c.line(margin, y + 12, table_right, y + 12)
            c.drawString(col_no_x, y, "No")
            c.drawString(col_name_x, y, "Nama Kegiatan")
            c.drawString(col_date_x, y, "Tanggal")
            c.drawRightString(table_right, y, "Poin")
            # Bottom line
            c.setLineWidth(0.5)
            c.line(margin, y - 4, table_right, y - 4)
            return y - row_h

        def draw_table_row(y, no, name, date_str, point_val):
            """Draw a single row. Returns new y."""
            c.setFont("Times-Roman", 9)
            c.setFillColor(Color(0.1, 0.1, 0.1))
            c.drawString(col_no_x, y, str(no))
            # Truncate name if too long
            max_name_w = col_date_x - col_name_x - 10
            display_name = name
            while stringWidth(display_name, "Times-Roman", 9) > max_name_w and len(display_name) > 5:
                display_name = display_name[:-4] + "..."
            c.drawString(col_name_x, y, display_name)
            c.drawString(col_date_x, y, date_str)
            c.drawRightString(table_right, y, str(point_val))
            return y - row_h

        def draw_section(title, items, curr_y):
            """Draw a section title + table for a category. Handles page breaks."""
            if not items:
                return curr_y

            # Check if we have enough space for at least the title + header + 1 row
            needed = 50
            if curr_y - needed < page_bottom:
                hb = start_new_page()
                curr_y = hb - 20

            # Section title
            c.setFont("Times-Bold", 11)
            c.setFillColor(Color(0.2, 0.2, 0.2))
            c.drawString(margin, curr_y, title)
            curr_y -= 18

            # Table header
            curr_y = draw_table_header(curr_y)

            subtotal = 0
            for i, item in enumerate(items):
                # Check page break
                if curr_y - row_h < page_bottom:
                    # Draw a "lanjutan" note
                    c.setFont("Times-Italic", 8)
                    c.setFillColor(Color(0.5, 0.5, 0.5))
                    c.drawRightString(table_right, curr_y, "(lanjutan di halaman berikutnya)")
                    hb = start_new_page()
                    curr_y = hb - 20
                    # Re-draw section title + header
                    c.setFont("Times-Bold", 11)
                    c.setFillColor(Color(0.2, 0.2, 0.2))
                    c.drawString(margin, curr_y, f"{title} (lanjutan)")
                    curr_y -= 18
                    curr_y = draw_table_header(curr_y)

                pt = item.get('point', 0)
                subtotal += pt
                job_info = item.get('job', '')
                name_display = item.get('title', '-')
                if job_info and job_info != '-':
                    name_display = f"{name_display} ({job_info})"

                curr_y = draw_table_row(
                    curr_y,
                    i + 1,
                    name_display,
                    fmt_date(item.get('date')),
                    pt
                )

            # Subtotal line
            c.setLineWidth(0.3)
            c.line(margin, curr_y + row_h - 4, table_right, curr_y + row_h - 4)
            c.setFont("Times-Bold", 9)
            c.setFillColor(Color(0.1, 0.1, 0.1))
            c.drawString(col_date_x, curr_y, "Subtotal")
            c.drawRightString(table_right, curr_y, str(subtotal))
            curr_y -= row_h + 8

            return curr_y

        # Draw each section
        sections = [
            ("A. Kepanitiaan Agenda", activities_details.get('committees', [])),
            ("B. Kepesertaan Agenda", activities_details.get('participants', [])),
            ("C. Proyek", activities_details.get('projects', [])),
            ("D. Aspirasi", activities_details.get('aspirations', [])),
        ]

        for sec_title, sec_items in sections:
            curr_y = draw_section(sec_title, sec_items, curr_y)

        # ── Grand Total ──
        if curr_y - 30 < page_bottom:
            hb = start_new_page()
            curr_y = hb - 20

        c.setLineWidth(1)
        c.line(margin, curr_y, table_right, curr_y)
        curr_y -= 16
        c.setFont("Times-Bold", 12)
        c.setFillColor(Color(0.1, 0.1, 0.1))
        c.drawString(margin, curr_y, "TOTAL POIN KEAKTIFAN")
        c.drawRightString(table_right, curr_y, str(total_point))
        curr_y -= 6
        c.setLineWidth(1.5)
        c.line(margin, curr_y, table_right, curr_y)

        c.save()
        buffer.seek(0)

        # Upload
        filename = f"{member_data.get('NIM')}/SKA_{point_data.get('semester')}.pdf"
        r2_key = f"documents/activiness-letter/{filename}"
        public_url = upload_bytes_to_r2(buffer.getvalue(), "application/pdf", r2_key)

        return jsonify({
            "success": True,
            "url": public_url,
            "filename": filename,
            "signatureLocations": sig_locations
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def rupiah_format(angka, with_prefix=False, desimal=2):
    # Format number with comma as thousands separator and dot as decimal
    s = "{:,.{}f}".format(angka, desimal)
    # Swap separators: 1,234.56 -> 1.234,56
    rupiah = s.replace(',', 'v').replace('.', ',').replace('v', '.')
    if with_prefix:
        return "Rp. {}".format(rupiah)
    return rupiah
    
# 4. GENERATE TICKET
@app.route('/api/pdf/ticket', methods=['POST'])
def generate_ticket():
    try:
        body = request.json
        agenda = body.get('agenda')
        amount = body.get('amount')
        if amount is None:
            amount = 0
        participant = body.get('participant') or {}
        role = body.get('role', 'participant')

        # Ukuran Ticket (Landscape)
        width, height = 600, 250
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=(width, height))
        
        # --- CONFIG ---
        primary_color = Color(0, 141/255, 211/255) # Blue #008DD3
        secondary_color = Color(252/255, 132/255, 17/255) # Orange #FC8411
        text_color = Color(0.2, 0.2, 0.2)
        stub_x = 450
        
        # --- BACKGROUND & SHAPES ---
        # 1. Left Accent (Rounded)
        c.setFillColor(primary_color)
        p = c.beginPath()
        p.moveTo(0, 0)
        p.lineTo(40, 0)
        p.lineTo(40, height)
        p.lineTo(0, height)
        c.drawPath(p, fill=1, stroke=0)
        
        # Semi-circle cutout left
        c.setFillColor(white)
        c.circle(0, height/2, 15, fill=1, stroke=0)

        # 2. Right Stub Background
        c.setFillColor(primary_color)
        c.rect(stub_x, 0, width - stub_x, height, fill=1, stroke=0)
        
        # Semi-circle cutout divider
        c.setFillColor(white)
        c.circle(stub_x, height, 10, fill=1, stroke=0) # Top
        c.circle(stub_x, 0, 10, fill=1, stroke=0) # Bottom

        # dashed line
        c.setStrokeColor(white)
        c.setLineWidth(2)
        c.setDash(4, 4)
        c.line(stub_x, 20, stub_x, height - 20)
        c.setDash([]) # Reset

        # --- CONTENT: LEFT BAND ---
        c.saveState()
        c.translate(25, height/2)
        c.rotate(90)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(0, 0, "NO. " + str(participant.get('_id', '000000'))[-6:].upper())
        c.restoreState()

        # --- CONTENT: MAIN AREA ---
        c.setFillColor(text_color)
        
        # Logos Coordinates (Calculated early for width constraints)
        logo_size = 35
        logo_y = height - 45 - logo_size
        itsnu_x = stub_x - 30 - (logo_size * 2) - 10
        hima_x = stub_x - 30 - logo_size
        
        # Date String Logic (Calculated early for width constraints)
        date_info = agenda.get('date') or {}
        date_start = date_info.get('start') or '2025-01-01'
        date_end = date_info.get('end') or '2025-01-01'
        d_start = date_start.split('T')[0]
        d_end = date_end.split('T')[0]
        date_str = d_start if d_start == d_end else f"{d_start} to {d_end}"
        date_str_width = max(stringWidth("DATE", "Helvetica-Bold", 10), stringWidth(date_str, "Helvetica", 10))

        # Title
        title = agenda.get('title', 'AGENDA').upper()
        title_font_size = 22
        max_title_width = itsnu_x - 60 - 10
        
        while stringWidth(title, "Helvetica-Bold", title_font_size) > max_title_width and title_font_size > 12:
            title_font_size -= 1
            
        if stringWidth(title, "Helvetica-Bold", title_font_size) > max_title_width:
            while stringWidth(title + "...", "Helvetica-Bold", title_font_size) > max_title_width and len(title) > 5:
                title = title[:-1]
            title += "..."
            
        c.setFont("Helvetica-Bold", title_font_size)
        c.drawString(60, height - 55, title)
        
        c.setFont("Helvetica", 10)
        c.setFillColor(secondary_color)
        c.drawString(60, height - 70, "HIMATIKA EVENT TICKET")
        
        # Main Info
        c.setFillColor(text_color)
        member_dict = participant.get('member') or participant.get('guest') or {}
        member_name = member_dict.get('fullName', 'Peserta').upper()
        
        name_font_size = 14
        max_name_width = (stub_x - 30) - date_str_width - 60 - 20
        
        while stringWidth(member_name, "Helvetica-Bold", name_font_size) > max_name_width and name_font_size > 9:
            name_font_size -= 0.5
            
        if stringWidth(member_name, "Helvetica-Bold", name_font_size) > max_name_width:
            while stringWidth(member_name + "...", "Helvetica-Bold", name_font_size) > max_name_width and len(member_name) > 5:
                member_name = member_name[:-1]
            member_name += "..."
            
        c.setFont("Helvetica-Bold", name_font_size)
        c.drawString(60, height - 110, member_name)
        
        c.setFont("Helvetica", 10)
        c.setFillColor(gray)
        c.drawString(60, height - 125, role.upper())

        # Logos
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        logo_itsnu_path = os.path.join(base_dir, 'assets', 'img', 'itsnu-logo.png')
        logo_hima_path = os.path.join(base_dir, 'assets', 'img', 'logo.png')
        
        if os.path.exists(logo_itsnu_path):
            c.drawImage(ImageReader(logo_itsnu_path), itsnu_x, logo_y, width=logo_size, height=logo_size, mask='auto', preserveAspectRatio=True)
        if os.path.exists(logo_hima_path):
            c.drawImage(ImageReader(logo_hima_path), hima_x, logo_y, width=logo_size, height=logo_size, mask='auto', preserveAspectRatio=True)

        # Date Label
        c.setFillColor(secondary_color)
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(stub_x - 30, height - 105, "DATE")
        c.setFillColor(text_color)
        c.setFont("Helvetica", 10)
        c.drawRightString(stub_x - 30, height - 120, date_str)

        # Bottom Boxes (Price/Type)
        def draw_box(x, y, text, bg_color, box_width=110):
            c.setFillColor(bg_color)
            c.rect(x, y, box_width, 30, fill=1, stroke=0)
            c.setFillColor(white if bg_color == primary_color else text_color)
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(x + (box_width/2), y + 10, text)

        draw_box(60, 40, role.upper(), primary_color, box_width=110)
        draw_box(180, 40, rupiah_format(amount, True), secondary_color, box_width=110)

        # Sponsors (Bottom right before QR)
        sponsors = agenda.get('configuration', {}).get('sponsors', [])
        pdf_sponsors = [s for s in sponsors if s.get('showOnPdf')]
        pdf_sponsors = pdf_sponsors[:3]
        if pdf_sponsors:
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(gray)
            c.drawString(60, 25, "SUPPORTED BY:")
            
            sp_x = 135
            sp_y = 15
            sp_size = 20
            for sp in pdf_sponsors:
                logo_url = sp.get('logo')
                if logo_url:
                    try:
                        import urllib.request
                        req = urllib.request.Request(logo_url, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=5) as response:
                            img_data = response.read()
                            img_mem = io.BytesIO(img_data)
                            c.drawImage(ImageReader(img_mem), sp_x, sp_y, width=sp_size, height=sp_size, mask='auto', preserveAspectRatio=True)
                            sp_x += sp_size + 10
                    except Exception as e:
                        print(f"Failed to load sponsor logo: {e}")

        # QR Code (Middle right)
        qr_payload = {"id": participant.get('_id'), "role": role}
        qr = qrcode.QRCode(box_size=2, border=0)
        qr.add_data(json.dumps(qr_payload))
        qr.make(fit=True)
        img_qr_main = qr.make_image(fill_color="black", back_color="white")
        qr_mem = io.BytesIO()
        img_qr_main.save(qr_mem, format='PNG')
        qr_mem.seek(0)
        qr_img = ImageReader(qr_mem)
        
        c.drawImage(qr_img, stub_x - 120, 25, 90, 90)

        # --- CONTENT: RIGHT STUB ---
        c.saveState()
        c.translate(stub_x + 10, height/2)
        
        c.setFillColor(text_color)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(10, 40, "ADMIT ONE")
        
        c.setFont("Helvetica", 10)
        c.drawString(10, 10, (member_name[:15] + "...") if len(member_name) > 15 else member_name)
        

        
        c.restoreState()

        c.save()
        buffer.seek(0)
        
        # Return PDF directly
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{body.get('id', 'ticket')}.pdf"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# 5. TOOLS
@app.route('/api/tools/qr', methods=['POST'])
def generate_qr_tool():
    try:
        body = request.json
        text = body.get('text')
        
        if not text:
            return jsonify({"error": "No text provided"}), 400

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(text)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")
        
        return jsonify({
            "success": True,
            "dataUrl": f"data:image/png;base64,{img_str}"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/tools/compress-image', methods=['POST'])
def compress_image_tool():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        quality = int(request.form.get('quality', 80))
        max_width = request.form.get('maxWidth')
        max_height = request.form.get('maxHeight')
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Open image using Pillow
        img = Image.open(file.stream)
        
        # Convert to RGB if necessary (e.g. for JPEG saving)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
            
        # Resize if needed
        if max_width or max_height:
            w, h = img.size
            ratio = min(
                int(max_width)/w if max_width else 1, 
                int(max_height)/h if max_height else 1
            )
            # Only downscale
            if ratio < 1:
                new_size = (int(w*ratio), int(h*ratio))
                img = img.resize(new_size, Image.Resampling.LANCZOS)
                
        # Compress
        output_buffer = io.BytesIO()
        # Default to JPEG for compression efficiency unless original was transparency heavy, 
        # but here we force JPEG/WebP as requested or just standard JPEG for simplicity 
        # based on 'browser-image-compression' replacement context usually implying JPEG/WebP.
        # Let's return JPEG for now as it's most common for 'compression'.
        
        img.save(output_buffer, format='JPEG', quality=quality, optimize=True)
        output_buffer.seek(0)
        
        # Encode to base64 to return easily (or binary if preferred, but base64 is easier for standard JSON API)
        # However, `customReadMultipartFormData` in Nuxt expects a buffer. 
        # Sending base64 is safer for JSON response, or we can return raw bytes with correct mimetype.
        # Let's return base64 and decode in Nuxt to Buffer.
        
        img_str = base64.b64encode(output_buffer.getvalue()).decode("utf-8")
        
        return jsonify({
            "success": True,
            "data": img_str,
            "mime": "image/jpeg"
        })

    except Exception as e:
        print(f"Error compressing image: {e}")
        return jsonify({'error': str(e)}), 500





@app.route('/api/tools/upload-image', methods=['POST'])
def upload_image_to_r2():
    """Upload an image to R2 and return its public URL.
    Preserves PNG transparency — important for wet signature images.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        quality = int(request.form.get('quality', 95))
        img = Image.open(file.stream)
        out = io.BytesIO()

        # Preserve transparency for PNG signatures
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            img.convert('RGBA').save(out, format='PNG', optimize=True)
            content_type = 'image/png'
            ext = 'png'
        else:
            img.convert('RGB').save(out, format='JPEG', quality=quality, optimize=True)
            content_type = 'image/jpeg'
            ext = 'jpeg'

        out.seek(0)
        import uuid
        key = f"uploads/signatures/{uuid.uuid4().hex}.{ext}"
        public_url = upload_bytes_to_r2(out.getvalue(), content_type, key)
        return jsonify({'success': True, 'url': public_url})

    except Exception as e:
        print(f"upload_image_to_r2 error: {e}")
        return jsonify({'error': str(e)}), 500

# 5. CERTIFICATE PREVIEW

@app.route('/api/pdf/certificate-preview', methods=['POST'])
def certificate_preview():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            file.save(temp_pdf.name)
            temp_pdf_path = temp_pdf.name
        
        doc = fitz.open(temp_pdf_path)
        page = doc.load_page(0) # First page only
        
        # Capture dimensions before closing doc
        width = page.rect.width
        height = page.rect.height
        
        pix = page.get_pixmap(dpi=150) # Moderate quality for preview
        
        img_bytes = io.BytesIO(pix.tobytes("png"))
        
        doc.close()
        # os.remove(temp_pdf_path) # Keep for upload
        
        # Upload preview to R2
        filename = f"preview_{datetime.datetime.now().timestamp()}.png"
        r2_key = f"certificates/previews/{filename}"
        public_url = upload_bytes_to_r2(img_bytes.getvalue(), "image/png", r2_key)

        # Upload PDF Template to R2
        with open(temp_pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        
        pdf_filename = f"template_{datetime.datetime.now().timestamp()}.pdf"
        pdf_key = f"certificates/templates/{pdf_filename}"
        pdf_url = upload_bytes_to_r2(pdf_bytes, "application/pdf", pdf_key)
        
        os.remove(temp_pdf_path)

        return jsonify({
            "success": True,
            "url": public_url,
            "pdfUrl": pdf_url,
            "width": width,
            "height": height
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# 6. GENERATE CERTIFICATE
@app.route('/api/pdf/certificate', methods=['POST'])
def generate_certificate():
    try:
        body = request.json
        template_url = body.get('templateUrl')
        items = body.get('items', [])
        data = body.get('data', {}) # { name, role, ... }
        docNo = ''
        
        if not template_url:
            return jsonify({"error": "Missing template URL"}), 400

        # Download Template
        # If url is from R2, it is publicly accessible?
        # Assuming upload_bytes_to_r2 returns public url.
        response = requests.get(template_url)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch template"}), 400
        
        input_pdf_stream = io.BytesIO(response.content)
        reader = PdfReader(input_pdf_stream)
        writer = PdfWriter()
        
        page = reader.pages[0]
        p_w = float(page.mediabox.width)
        p_h = float(page.mediabox.height)
        
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(p_w, p_h))
        
        for item in items:
            itype = item.get('type')
            x = float(item.get('x', 0))
            y_top = float(item.get('y', 0)) # UI uses Top-Left
            w = float(item.get('width', 0))
            h = float(item.get('height', 0))
            
            # Convert Y to Bottom-Left
            y_bot = p_h - y_top - float(item.get('fontSize', 12)) 
            
            font_size = float(item.get('fontSize', 12))
            font_name = item.get('fontFamily', 'Times-Roman')
            font_weight = item.get('fontWeight', 'normal')
            
            if font_weight == 'bold' and font_name == 'Times-Roman':
                font_name = 'Times-Bold'
            
            align = item.get('align', 'left')
            color_hex = item.get('color', '#000000')
            
            c.setFont(font_name, font_size)
            c.setFillColor(Color(
                int(color_hex[1:3], 16)/255.0,
                int(color_hex[3:5], 16)/255.0,
                int(color_hex[5:7], 16)/255.0
            ))
            
            # Align Logic
            # Frontend X, Y is Top-Left of the bounding box.
            # ReportLab Y is Bottom-Left.
            
            # For Text:
            # If align 'left': drawString at x, y_bot.
            # If align 'center': drawCentredString at x + w/2, y_bot.
            # If align 'right': drawRightString at x + w, y_bot.
            
            # Y Correction:
            # We want the text to be vertically centered in the box or at least reasonably placed.
            # y_top is the top of the box.
            # h is the height of the box.
            # font_size is the height of text approx.
            # Let's target the baseline.
            # y_bot = p_h - (y_top + h/2 - font_size/3) # Rough vertical center
            
            # Better: Let's assume the user positions the box where they want the text.
            # Standard PDF text is drawn from baseline.
            # We want visual vertical centering.
            # Center of Box = p_h - (y_top + h/2)
            # Baseline should be slightly below center.
            # y_bot = Center - (font_size / 3)
            
            y_bot = (p_h - (y_top + h/2)) - (font_size / 2.5)

            content = ""
            if itype == 'name':
                content = data.get('name', '')
            elif itype == 'role':
                content = data.get('role', '')
            elif itype == 'text':
                content = item.get('value', '')
            elif itype == 'date':
                content = data.get('date', datetime.datetime.now().strftime("%d %B %Y"))
            elif itype == 'code':
                content = item.get('value', '')
                docNo = content
                

                
            # Draw Logic
            if itype == 'qr':
                content = data.get('qr_data', 'https://himatika.org')
                qr_item = qrcode.QRCode(border=0)
                qr_item.add_data(content)
                qr_item.make(fit=True)
                qr_img_pil = qr_item.make_image(fill_color="black", back_color="white")
                qr_img_bytes = io.BytesIO()
                qr_img_pil.save(qr_img_bytes, format='PNG')
                qr_img_bytes.seek(0)
                y_bot_qr = p_h - (y_top + h)
                c.drawImage(ImageReader(qr_img_bytes), x, y_bot_qr, width=w, height=h, mask='auto', preserveAspectRatio=True)
                continue

            elif itype == 'signature':
                signer_type = item.get('signerType', 'external')
                sig_img_url = item.get('value', '')       # URL gambar TTD (eksternal)
                signer_name = item.get('signerName', '')  # nama cetak (eksternal)
                signer_as   = item.get('signerAs', '')

                sig_img_reader = None
                if signer_type == 'external' and sig_img_url:
                    try:
                        img_resp = requests.get(sig_img_url, timeout=10)
                        if img_resp.status_code == 200:
                            sig_img_reader = ImageReader(io.BytesIO(img_resp.content))
                    except Exception as se:
                        print(f"Signature image fetch error: {se}")

                # Mode sistem: area tengah dibiarkan kosong → QR di-overlay nanti
                display_name = signer_name if signer_type == 'external' else ''

                draw_signature_box(c, x, y_top, w, h, p_h,
                                   sig_name=display_name,
                                   sig_as=signer_as,
                                   sig_img=sig_img_reader,
                                   overlap=sig_img_reader is not None)  # hanya TTD basah yang tumpang tindih
                continue

            # Text Drawing with Align
            if align == 'center':
                c.drawCentredString(x + w/2, y_bot, content)
            elif align == 'right':
                c.drawRightString(x + w, y_bot, content)
            else:
                c.drawString(x, y_bot, content)

        
        c.save()
        packet.seek(0)
        
        page.merge_page(PdfReader(packet).pages[0])
        writer.add_page(page)
        
        out_buffer = io.BytesIO()
        writer.write(out_buffer)
        out_buffer.seek(0)
        filename = f"Certificate_{docNo}_{data.get('name', 'Unknown')}.pdf"
        r2_key = f"certificates/generated/{filename}"
        public_url = upload_bytes_to_r2(out_buffer.getvalue(), "application/pdf", r2_key)
        
        return jsonify({
            "success": True,
            "url": public_url
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- MEDIA PROCESSING ---

def _compress_video_task(file_key: str, video_id: str, callback_url: str) -> None:
    """
    Background task: download raw video from R2, compress with FFmpeg, re-upload, notify backend.
    Uses context manager for temp files to prevent memory leaks.
    """
    s3 = get_s3_client()
    bucket_name = os.environ.get('R2_BUCKET_NAME', '')
    public_domain = os.environ.get('R2_PUBLIC_DOMAIN', '').rstrip('/')

    input_path = None
    output_path = None

    try:
        # 1. Extract extension from file_key if possible
        ext = os.path.splitext(file_key)[1]
        if not ext:
            ext = ".mp4"  # Default fallback

        # Download raw video from R2
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_input:
            input_path = tmp_input.name
            s3.download_fileobj(bucket_name, file_key, tmp_input)

        # 2. Compress with FFmpeg: Force output to WebM format
        output_path = input_path + "_compressed.webm"
        ffmpeg_cmd = [
            "ffmpeg", "-y",
            "-i", input_path,
            "-vf", "scale=-2:720",
            "-c:v", "libvpx-vp9",
            "-crf", "30",
            "-b:v", "0",
            "-c:a", "libopus",
            "-b:a", "128k",
            "-deadline", "realtime",
            "-cpu-used", "4",
            output_path
        ]
        content_type = "video/webm"
        base_key, _ = os.path.splitext(file_key)
        compressed_key = base_key.replace("raw_", "compressed_", 1) + ".webm"

        result = subprocess.run(ffmpeg_cmd, capture_output=True, text=True, timeout=600)
        if result.returncode != 0:
            print(f"FFmpeg stderr: {result.stderr}")
            raise RuntimeError(f"FFmpeg failed with code {result.returncode}")

        # 3. Upload compressed video to R2
        with open(output_path, "rb") as f:
            compressed_bytes = f.read()

        processed_url = upload_bytes_to_r2(compressed_bytes, content_type, compressed_key)

        # 4. Delete raw video from R2
        try:
            s3.delete_object(Bucket=bucket_name, Key=file_key)
        except Exception as del_err:
            print(f"Warning: Failed to delete raw video {file_key}: {del_err}")

        # 5. Notify backend via webhook
        requests.post(
            callback_url,
            json={
                "videoId": video_id,
                "status": "completed",
                "processedUrl": processed_url,
            },
            timeout=30,
        )
        print(f"Video {video_id} compressed successfully: {processed_url}")

    except Exception as e:
        print(f"Video compression failed for {video_id}: {e}")
        import traceback
        traceback.print_exc()

        try:
            requests.post(
                callback_url,
                json={
                    "videoId": video_id,
                    "status": "failed",
                },
                timeout=30,
            )
        except Exception as notify_err:
            print(f"Failed to notify backend of failure: {notify_err}")

    finally:
        for path in [input_path, output_path]:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass


@app.route('/api/media/compress-video', methods=['POST'])
def compress_video():
    try:
        body = request.json
        file_key: str = body.get('fileKey', '')
        video_id: str = body.get('videoId', '')
        callback_url: str = body.get('callbackUrl', '')

        if not all([file_key, video_id, callback_url]):
            return jsonify({"error": "Missing required parameters"}), 400

        thread = threading.Thread(
            target=_compress_video_task,
            args=(file_key, video_id, callback_url),
            daemon=True,
        )
        thread.start()

        return jsonify({
            "success": True,
            "message": "Video compression started",
            "videoId": video_id,
        }), 202

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    from waitress import serve
    serve(app, host="0.0.0.0", port=8000)
