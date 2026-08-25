import sys
import os
import io
import base64
import tempfile
import datetime
import requests
import json
import fitz
from pypdf import PdfReader, PdfWriter
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import black, white, gray, Color
from reportlab.pdfbase.pdfmetrics import stringWidth
from openpyxl import Workbook, load_workbook
import threading
import subprocess
from flask import Blueprint, request, send_file, jsonify
from utils.db import get_members_collection
from utils.storage import upload_bytes_to_r2, get_s3_client
from utils.helpers import flatten_object

# We will need the helper functions here or in utils/helpers.py
# For now, let's just dump them if needed, or better, we moved them to utils/helpers.py in another branch, but we are on main branch.
# Let's extract the helpers from index.py (lines 46 - 165)

bp = Blueprint('sheet', __name__)

def month_to_roman(date_obj):
    month = date_obj.month
    roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
    return roman_numerals[month - 1]

def format_date_indo(date_obj):
    months = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]
    day = date_obj.day
    month = months[date_obj.month - 1]
    year = date_obj.year
    return f"{day} {month} {year}"

def draw_wrapped_text(canvas_obj, text, x, y, max_width, font_name, font_size, line_height=None):
    """
    Menggambar teks yang otomatis turun baris jika melebihi max_width.
    """
    if line_height is None:
        line_height = font_size + 4

    words = text.split(" ")
    current_line = []
    current_y = y
    
    canvas_obj.setFont(font_name, font_size)

    for word in words:
        test_line = " ".join(current_line + [word])
        width = stringWidth(test_line, font_name, font_size)
        
        if width <= max_width:
            current_line.append(word)
        else:
            canvas_obj.drawString(x, current_y, " ".join(current_line))
            current_line = [word]
            current_y -= line_height

    if current_line:
        canvas_obj.drawString(x, current_y, " ".join(current_line))
    
    return current_y


def draw_signature_box(c, x, y_top, w, h, p_h,
                        sig_name='', sig_as='', sig_img=None, overlap=False):
    """
    Layout kotak tanda tangan:
        ┌─────────────────┐
        │  Jabatan (atas) │  ← 15% h, font min 12pt
        ├─────────────────┤
        │  [TTD / QR]     │  ← gambar besar, ujung bawahnya
        │    ↕ overlap    │    menumpangi area nama (~40%)
        ├─────────────────┤  ← garis tipis
        │  Nama (bawah)   │  ← 20% h, font min 12pt
        └─────────────────┘
    Gambar digambar TERAKHIR sehingga secara visual
    "tinta tanda tangan" berada di atas teks nama.
    sig_img : ImageReader-compatible, atau None
    """
    role_h   = h * 0.15   # jabatan di atas
    name_h   = h * 0.20   # nama di bawah
    img_h    = h - role_h - name_h   # sisa → gambar

    # Koordinat ReportLab (bottom-left origin)
    y_bot    = p_h - (y_top + h)
    name_y0  = y_bot               # bawah area nama
    name_y1  = y_bot + name_h      # batas atas nama / batas bawah garis
    img_y0   = name_y1             # bawah area gambar (tanpa overlap)
    img_y1   = img_y0 + img_h      # atas area gambar / batas bawah jabatan
    role_y0  = img_y1

    c.saveState()

    # ── 1. Jabatan teks (atas) ─────────────────────────────────
    if sig_as:
        role_fs = max(12, min(14, role_h * 0.60))
        c.setFont('Helvetica', role_fs)
        c.setFillColor(Color(0.2, 0.2, 0.2))
        role_text_y = role_y0 + (role_h - role_fs) / 2
        c.drawCentredString(x + w / 2, role_text_y, sig_as[:50])

    # ── 2. Garis atas (antara jabatan & gambar) ────────────────
    c.setStrokeColor(Color(0.5, 0.5, 0.5))
    c.setLineWidth(0.3)
    c.line(x, img_y1, x + w, img_y1)

    # ── 3. Garis bawah (antara gambar & nama) ─────────────────
    c.setStrokeColor(Color(0.5, 0.5, 0.5))
    c.setLineWidth(0.5)
    c.line(x, name_y1, x + w, name_y1)

    # ── 4. Nama teks (bawah, di bawah garis) ──────────────────
    if sig_name:
        name_fs = max(12, min(14, name_h * 0.60))
        c.setFont('Helvetica-Bold', name_fs)
        c.setFillColor(Color(0.1, 0.1, 0.1))
        name_text_y = name_y0 + (name_h - name_fs) / 2
        c.drawCentredString(x + w / 2, name_text_y, sig_name[:40])

    # ── 5. Gambar TTD / QR (digambar TERAKHIR → di atas nama) ──
    if sig_img is not None:
        if overlap:
            # Wet signature: gambar tumpang-tindih ke area nama (50% name_h)
            overlap_px  = name_h * 0.50
            draw_y      = img_y0 - overlap_px
            draw_h      = img_h + overlap_px
        else:
            # QR: tetap di dalam area gambar, tidak tumpang tindih
            draw_y  = img_y0
            draw_h  = img_h
        side     = min(w * 0.92, draw_h)
        draw_x   = x + (w - side) / 2
        center_y = draw_y + (draw_h - side) / 2
        c.drawImage(sig_img, draw_x, center_y,
                    width=side, height=side,
                    mask='auto', preserveAspectRatio=False)

    c.restoreState()


@bp.route('/api/sheet/import', methods=['POST'])
def import_generic_sheet():
    try:
        if 'file' not in request.files:
             return jsonify({"error": "No file part"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active # Default to first sheet or active one
        
        data = []
        headers = []
        
        # Read headers (Row 1)
        header_row_gen = ws.iter_rows(min_row=1, max_row=1)
        try:
            header_row = next(header_row_gen)
            for cell in header_row:
                headers.append(cell.value)
        except StopIteration:
            pass # Empty sheet
            
        # Read data (Row 2 onwards)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
                
            row_data = {}
            for i, cell_value in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    if header is None: continue

                    # Handle nested keys (e.g. "user.name")
                    keys = str(header).split('.')
                    current_ref = row_data
                    
                    for k_idx, key in enumerate(keys):
                        if k_idx == len(keys) - 1:
                            # Last key, assign value
                            current_ref[key] = cell_value
                        else:
                            # Create nested dict if not exists
                            if key not in current_ref:
                                current_ref[key] = {}
                            current_ref = current_ref[key]
                            
            data.append(row_data)
            
        return jsonify({
            "success": True,
            "data": data
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# 1. GENERIC EXCEL EXPORT
@bp.route('/api/sheet/export', methods=['POST'])
def export_generic_sheet():
    try:
        body = request.json
        complex_data = body.get('data', [])
        title = body.get('title', 'Export')
        headers_input = body.get('headers', None)

        wb = Workbook()
        ws = wb.active
        ws.title = title[:30] # Excel sheet limit

        if not complex_data:
            return jsonify({"message": "No data"}), 400

        if headers_input:
            headers = headers_input
        else:
            all_headers = set()
            for item in complex_data:
                flat = flatten_object(item)
                all_headers.update(flat.keys())
            headers = list(all_headers)
            
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for item in complex_data:
            flat = flatten_object(item)
            row = [str(flat.get(h, "")) if isinstance(flat.get(h, ""), (list, dict)) else flat.get(h, "") for h in headers]
            ws.append(row)

        # Style data cells
        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_alignment

        # Auto adjust column width
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 60) # Max width to avoid overly wide columns

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{title}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

